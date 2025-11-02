# -*- coding: utf-8 -*-
"""
Private RAG API Server - Enterprise Grade
- Fabrix-inspired features with RAG and Standard LLM modes
- Maximum stability and error recovery
"""
import os
import io
import time
import json
import logging
import sqlite3
import traceback
from pathlib import Path
from typing import List, Optional, Dict, Any, Iterable, AsyncGenerator
from datetime import datetime
import uuid
import shutil
import asyncio
from contextlib import asynccontextmanager

import httpx
from fastapi import FastAPI, UploadFile, File, HTTPException, Query, status, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel, Field, validator

from qdrant_client import QdrantClient
from qdrant_client.http.models import (
    Distance, VectorParams, Filter, FieldCondition, MatchValue, FilterSelector
)
from fastembed import TextEmbedding

# ==================== 설정 ====================
QDRANT_LOCAL = os.getenv("QDRANT_LOCAL", "1") == "1"
QDRANT_HOST = os.getenv("QDRANT_HOST", "localhost")
QDRANT_PORT = int(os.getenv("QDRANT_PORT", "6333"))
COLLECTION_NAME = os.getenv("COLLECTION_NAME", "documents")

OLLAMA_HOST = os.getenv("OLLAMA_HOST", "localhost")
OLLAMA_PORT = int(os.getenv("OLLAMA_PORT", "11434"))
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.1:8b")
OLLAMA_TIMEOUT = int(os.getenv("OLLAMA_TIMEOUT", "180"))

CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "600"))
CHUNK_OVERLAP = int(os.getenv("CHUNK_OVERLAP", "250"))
FASTEMBED_BATCH = int(os.getenv("FASTEMBED_BATCH", "256"))
UPSERT_BATCH = int(os.getenv("UPSERT_BATCH", "5000"))

# 경로 설정
BASE_DIR = Path(__file__).parent.resolve()
PROJECT_ROOT = BASE_DIR.parent
DEFAULT_DATA_DIR = (PROJECT_ROOT / "data").resolve()
EXPORT_DIR = (PROJECT_ROOT / "exports").resolve()
QDRANT_DIR = (PROJECT_ROOT / "qdrant_storage").resolve()
CHAT_HISTORY_DIR = (PROJECT_ROOT / "chat_history").resolve()

CURRENT_DATA_DIR: Path = DEFAULT_DATA_DIR

for d in (DEFAULT_DATA_DIR, EXPORT_DIR, QDRANT_DIR, CHAT_HISTORY_DIR):
    d.mkdir(parents=True, exist_ok=True)

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(BASE_DIR / "app.log", encoding="utf-8")
    ]
)
log = logging.getLogger("private-rag")

# ==================== 전역 변수 ====================
_qdrant: Optional[QdrantClient] = None
_embedding_model = None
_ollama_available = False
_available_models = []

# ==================== Lifespan ====================
@asynccontextmanager
async def lifespan(app: FastAPI):
    log.info("🚀 Starting Private RAG API Server (Fabrix Edition)...")
    try:
        await check_ollama_connection()
        await load_available_models()
        await initialize_embedding_model()
        log.info("✅ All systems initialized successfully")
    except Exception as e:
        log.error(f"⚠️ Initialization warning: {e}")

    yield

    log.info("🛑 Shutting down Private RAG API Server...")

# ==================== FastAPI ====================
app = FastAPI(
    title="Private RAG API - Fabrix Edition",
    version="3.0.0",
    description="Enterprise-Grade Private RAG System with Fabrix-style UI",
    lifespan=lifespan
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR = (PROJECT_ROOT / "frontend").resolve()
if FRONTEND_DIR.exists():
    app.mount("/ui", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="ui")

@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    log.error(f"Unhandled exception: {exc}\n{traceback.format_exc()}")
    return JSONResponse(
        status_code=500,
        content={
            "error": "Internal server error",
            "message": str(exc),
            "type": type(exc).__name__
        }
    )

@app.get("/")
def root():
    index = FRONTEND_DIR / "index.html"
    if index.exists():
        return FileResponse(index, headers={"Cache-Control": "no-store"})
    return {"status": "ok", "message": "Private RAG API Server - Fabrix Edition", "version": "3.0.0"}

# ==================== 헬퍼 함수 ====================
async def check_ollama_connection(retries: int = 3, delay: float = 1.0) -> bool:
    """Check Ollama connection with retry logic"""
    global _ollama_available
    for attempt in range(retries):
        try:
            async with httpx.AsyncClient(timeout=5.0) as client:
                r = await client.get(f"http://{OLLAMA_HOST}:{OLLAMA_PORT}/api/tags")
                _ollama_available = r.status_code == 200
                if _ollama_available:
                    log.info("✅ Ollama connection successful")
                return _ollama_available
        except Exception as e:
            if attempt < retries - 1:
                log.warning(f"Ollama connection attempt {attempt + 1}/{retries} failed: {e}")
                await asyncio.sleep(delay * (attempt + 1))  # Exponential backoff
            else:
                log.warning(f"Ollama not available after {retries} attempts: {e}")
                _ollama_available = False
                return False

async def load_available_models(retries: int = 3):
    """Load available Ollama models with retry logic"""
    global _available_models
    for attempt in range(retries):
        try:
            async with httpx.AsyncClient(timeout=5.0) as client:
                r = await client.get(f"http://{OLLAMA_HOST}:{OLLAMA_PORT}/api/tags")
                if r.status_code == 200:
                    data = r.json()
                    _available_models = [m["name"] for m in data.get("models", [])]
                    log.info(f"✅ Available models: {_available_models}")
                    return
        except Exception as e:
            if attempt < retries - 1:
                log.warning(f"Model load attempt {attempt + 1}/{retries} failed: {e}")
                await asyncio.sleep(1.0 * (attempt + 1))
            else:
                log.warning(f"Failed to load models after {retries} attempts: {e}")
                _available_models = []

async def initialize_embedding_model():
    global _embedding_model
    if _embedding_model is None:
        try:
            log.info("Loading embedding model...")
            _embedding_model = TextEmbedding(
                model_name="sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
            )
            log.info("✅ Embedding model loaded successfully")
        except Exception as e:
            log.error(f"❌ Failed to load embedding model: {e}")
            raise

def get_embedding_model():
    if _embedding_model is None:
        raise HTTPException(
            status_code=503,
            detail="Embedding model not initialized. Please restart the server."
        )
    return _embedding_model

def get_qdrant() -> QdrantClient:
    global _qdrant
    if _qdrant is not None:
        return _qdrant

    try:
        if QDRANT_LOCAL:
            log.info(f"Using Qdrant Embedded (path={QDRANT_DIR})")
            _qdrant = QdrantClient(path=str(QDRANT_DIR))
        else:
            log.info(f"Using Qdrant Remote ({QDRANT_HOST}:{QDRANT_PORT})")
            _qdrant = QdrantClient(host=QDRANT_HOST, port=QDRANT_PORT)
        return _qdrant
    except Exception as e:
        log.error(f"Failed to initialize Qdrant: {e}")
        raise HTTPException(
            status_code=503,
            detail=f"Cannot connect to Qdrant: {str(e)}"
        )

def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> Iterable[str]:
    if not text or not text.strip():
        return
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    n = len(text)
    if chunk_size <= 0 or n <= chunk_size:
        yield text.strip()
        return
    start = 0
    while start < n:
        end = min(start + chunk_size, n)
        piece = text[start:end].strip()
        if piece:
            yield piece
        if end == n:
            break
        start = end - overlap if overlap < chunk_size else end

def read_file_text(fp: Path) -> str:
    suf = fp.suffix.lower()
    log.info(f"Reading file: {fp.name} (type: {suf})")

    try:
        if suf in {".txt", ".md"}:
            data = fp.read_bytes()
            for enc in ("utf-8", "cp949", "euc-kr"):
                try:
                    text = data.decode(enc)
                    log.info(f"Successfully read {fp.name} with {enc} encoding ({len(text)} chars)")
                    return text
                except UnicodeDecodeError:
                    continue
            text = data.decode("utf-8", errors="ignore")
            log.info(f"Read {fp.name} with UTF-8 (ignore errors) ({len(text)} chars)")
            return text

        if suf == ".pdf":
            try:
                import fitz
                doc = fitz.open(str(fp))
                texts = [page.get_text("text") for page in doc]
                doc.close()
                result = "\n".join(texts)
                log.info(f"Successfully read PDF {fp.name} ({len(result)} chars)")
                return result
            except Exception as e1:
                log.warning(f"PyMuPDF failed for {fp.name}: {e1}, trying pypdf")
                from pypdf import PdfReader
                reader = PdfReader(str(fp))
                result = "\n".join(page.extract_text() or "" for page in reader.pages)
                log.info(f"Successfully read PDF {fp.name} with pypdf ({len(result)} chars)")
                return result

        if suf == ".docx":
            import docx2txt
            result = docx2txt.process(str(fp)) or ""
            log.info(f"Successfully read DOCX {fp.name} ({len(result)} chars)")
            return result

        if suf == ".csv":
            import csv
            data = fp.read_bytes()
            text = None
            for enc in ("utf-8", "cp949", "euc-kr"):
                try:
                    text = data.decode(enc)
                    break
                except UnicodeDecodeError:
                    pass
            if text is None:
                text = data.decode("utf-8", errors="ignore")
            sio = io.StringIO(text)
            reader = csv.reader(sio)
            result = "\n".join("\t".join("" if c is None else str(c) for c in row) for row in reader)
            log.info(f"Successfully read CSV {fp.name} ({len(result)} chars)")
            return result

        if suf == ".xlsx":
            log.info(f"Attempting to read XLSX file: {fp}")
            try:
                from openpyxl import load_workbook
                log.info(f"Loading workbook: {fp}")
                wb = load_workbook(filename=str(fp), read_only=True, data_only=True)
                log.info(f"Workbook loaded, sheets: {[ws.title for ws in wb.worksheets]}")

                parts = []
                for ws in wb.worksheets:
                    log.info(f"Processing sheet: {ws.title}")
                    parts.append(f"# Sheet: {ws.title}")
                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        row_text = "\t".join("" if c is None else str(c) for c in row)
                        if row_text.strip():  # 빈 행 제외
                            parts.append(row_text)
                            row_count += 1
                    log.info(f"Sheet {ws.title}: {row_count} rows extracted")

                wb.close()
                result = "\n".join(parts)
                log.info(f"Successfully read XLSX {fp.name} ({len(result)} chars, {len(parts)} lines)")
                return result
            except Exception as xlsx_err:
                log.error(f"XLSX reading failed for {fp.name}: {xlsx_err}")
                import traceback
                log.error(traceback.format_exc())
                return ""

        log.warning(f"Unsupported file type: {suf}")
        return ""

    except Exception as e:
        log.error(f"Error reading file {fp.name}: {e}")
        import traceback
        log.error(traceback.format_exc())
        return ""

def ensure_collection(vec_dim: int):
    try:
        client = get_qdrant()
        collections = [c.name for c in client.get_collections().collections]
        if COLLECTION_NAME not in collections:
            log.info(f"Creating collection '{COLLECTION_NAME}' (dim={vec_dim})")
            client.recreate_collection(
                collection_name=COLLECTION_NAME,
                vectors_config=VectorParams(size=vec_dim, distance=Distance.COSINE),
            )
    except Exception as e:
        log.error(f"Failed to ensure collection: {e}")
        raise HTTPException(status_code=500, detail=f"Collection creation failed: {str(e)}")

def embed_texts_batch(texts: List[str]) -> List[List[float]]:
    try:
        em = get_embedding_model()
        return [vec for vec in em.embed(texts, batch_size=FASTEMBED_BATCH)]
    except Exception as e:
        log.error(f"Embedding failed: {e}")
        raise HTTPException(status_code=500, detail=f"Embedding failed: {str(e)}")

async def call_ollama_stream(prompt: str, model: str = OLLAMA_MODEL, system: str = None, retries: int = 2) -> AsyncGenerator[str, None]:
    """Call Ollama with streaming response and retry logic"""
    if not _ollama_available:
        # Attempt to reconnect
        log.warning("Ollama not available, attempting to reconnect...")
        if not await check_ollama_connection(retries=2, delay=0.5):
            raise HTTPException(status_code=503, detail="Ollama is not available. Please ensure Ollama is running.")

    url = f"http://{OLLAMA_HOST}:{OLLAMA_PORT}/api/generate"
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": True
    }
    if system:
        payload["system"] = system

    last_error = None
    for attempt in range(retries + 1):
        try:
            async with httpx.AsyncClient(timeout=OLLAMA_TIMEOUT) as client:
                async with client.stream("POST", url, json=payload) as resp:
                    if resp.status_code != 200:
                        text = await resp.aread()
                        error_msg = text.decode(errors='ignore')

                        # Check if model not found
                        if "model" in error_msg.lower() and "not found" in error_msg.lower():
                            raise HTTPException(
                                status_code=404,
                                detail=f"Model '{model}' not found. Please pull it first: ollama pull {model}"
                            )

                        raise HTTPException(status_code=502, detail=f"Ollama error: {error_msg}")

                    async for line in resp.aiter_lines():
                        if not line:
                            continue
                        try:
                            obj = json.loads(line)
                            if obj.get("response"):
                                yield obj["response"]
                            if obj.get("done"):
                                return  # Successful completion
                        except json.JSONDecodeError:
                            continue
                    return  # Stream ended successfully

        except httpx.TimeoutException as e:
            last_error = e
            if attempt < retries:
                log.warning(f"Ollama timeout, retry {attempt + 1}/{retries}")
                await asyncio.sleep(1.0)
                continue
            raise HTTPException(status_code=504, detail="Ollama request timeout. The model may be too slow or the prompt too long.")

        except httpx.ConnectError as e:
            last_error = e
            if attempt < retries:
                log.warning(f"Ollama connection error, retry {attempt + 1}/{retries}")
                await check_ollama_connection(retries=1, delay=0.5)  # Try to reconnect
                await asyncio.sleep(1.0)
                continue
            raise HTTPException(status_code=503, detail="Cannot connect to Ollama. Please ensure Ollama is running.")

        except HTTPException:
            raise  # Re-raise HTTP exceptions without retry

        except Exception as e:
            last_error = e
            if attempt < retries:
                log.warning(f"Ollama error, retry {attempt + 1}/{retries}: {e}")
                await asyncio.sleep(1.0)
                continue
            log.error(f"Ollama stream error after {retries + 1} attempts: {e}")
            raise HTTPException(status_code=502, detail=f"Ollama connection failed: {str(e)}")

    # If we get here, all retries failed
    if last_error:
        raise HTTPException(status_code=502, detail=f"Failed after {retries + 1} attempts: {str(last_error)}")

def build_rag_prompt(question: str, contexts: List[Dict[str, Any]], language: str = "ko") -> str:
    sources_text = "\n\n".join([
        f"[{i+1}] {(c.get('payload') or {}).get('filename', 'unknown')}\n{(c.get('payload') or {}).get('text', '')[:1500]}"
        for i, c in enumerate(contexts)
    ])

    if language == "ko":
        system = (
            "당신은 문서 기반 질문에 답하는 AI 어시스턴트입니다. "
            "주어진 문서 컨텍스트만을 사용하여 정확하고 상세하게 답변하세요. "
            "정보가 불확실하면 모른다고 말하세요."
        )
    else:
        system = (
            "You are an AI assistant that answers questions based on documents. "
            "Use ONLY the provided context to answer accurately and in detail. "
            "If unsure, say you don't know."
        )

    return f"""{system}

Question: {question}

Context:
{sources_text}

Answer in {language}:"""

# ==================== 스키마 ====================
class IngestResponse(BaseModel):
    files_indexed: int
    chunks_indexed: int
    status: str = "success"

class QueryRequest(BaseModel):
    question: str = Field(..., min_length=1, max_length=5000)
    mode: str = Field(default="rag", pattern="^(rag|llm)$")  # RAG or Standard LLM
    top_k: int = Field(default=5, ge=1, le=20)
    language: str = Field(default="ko", pattern="^(ko|en)$")
    model: Optional[str] = None
    temperature: Optional[float] = Field(default=0.7, ge=0, le=2)
    system_prompt: Optional[str] = None
    selected_sources: Optional[List[str]] = None  # 선택된 문서 소스 목록

    @validator('question')
    def question_not_empty(cls, v):
        if not v.strip():
            raise ValueError('Question cannot be empty')
        return v.strip()

class StorageConfig(BaseModel):
    path: str = Field(..., min_length=1)

class ConversationExport(BaseModel):
    conversation_id: str
    title: str
    messages: List[Dict[str, Any]]
    created_at: str
    mode: str

# ==================== API 엔드포인트 ====================
@app.get("/health")
async def health():
    """Comprehensive health check with auto-recovery attempts"""
    try:
        # Check Qdrant
        qdrant_ok = False
        qdrant_error = None
        try:
            client = get_qdrant()
            client.get_collections()
            qdrant_ok = True
        except Exception as e:
            qdrant_error = str(e)
            log.warning(f"Qdrant health check failed: {e}")

        # Check Ollama with retry
        ollama_ok = await check_ollama_connection(retries=2, delay=0.5)

        # Reload models if Ollama reconnected
        if ollama_ok and not _available_models:
            await load_available_models(retries=1)

        # Check embedding model
        embedding_ok = _embedding_model is not None
        if not embedding_ok:
            try:
                await initialize_embedding_model()
                embedding_ok = True
            except:
                pass

        status = "healthy" if (qdrant_ok and ollama_ok and embedding_ok) else "degraded"
        if not ollama_ok:
            status = "critical"

        return {
            "status": status,
            "services": {
                "qdrant": {
                    "available": qdrant_ok,
                    "error": qdrant_error if not qdrant_ok else None
                },
                "ollama": {
                    "available": ollama_ok,
                    "models_loaded": len(_available_models),
                    "default_model": OLLAMA_MODEL
                },
                "embedding": {
                    "available": embedding_ok
                }
            },
            "available_models": _available_models,
            "timestamp": datetime.utcnow().isoformat()
        }
    except Exception as e:
        log.error(f"Health check failed: {e}")
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "error": str(e),
                "timestamp": datetime.utcnow().isoformat()
            }
        )

@app.get("/models")
async def get_models():
    """Get list of available Ollama models"""
    try:
        if not _ollama_available:
            await check_ollama_connection()

        if not _available_models:
            await load_available_models()

        return {"models": _available_models}
    except Exception as e:
        log.error(f"Failed to get models: {e}")
        return {"models": []}

# ==================== API v1 Routes ====================
@app.get("/api/v1/models")
async def get_models_v1():
    """Get list of available Ollama models (API v1)"""
    return await get_models()

@app.post("/api/v1/query")
async def query_v1(req: QueryRequest = Body(...)):
    """Query with RAG or LLM mode (API v1) - Non-streaming"""
    try:
        client = get_qdrant()
        model = req.model or OLLAMA_MODEL

        if req.mode == "rag":
            collections = [c.name for c in client.get_collections().collections]
            if COLLECTION_NAME not in collections:
                raise HTTPException(status_code=404, detail="No documents indexed")

            qvec = embed_texts_batch([req.question])[0]

            search_params = {
                "collection_name": COLLECTION_NAME,
                "query_vector": qvec,
                "limit": req.top_k,
            }

            if req.selected_sources and len(req.selected_sources) > 0:
                from qdrant_client.models import Filter, FieldCondition, MatchAny
                search_params["query_filter"] = Filter(
                    must=[
                        FieldCondition(
                            key="source",
                            match=MatchAny(any=req.selected_sources)
                        )
                    ]
                )

            hits = client.search(**search_params)
            contexts = [{"payload": h.payload, "score": float(h.score)} for h in hits]
            prompt = build_rag_prompt(req.question, contexts, req.language)

            # Collect streaming response
            answer = ""
            async for token in call_ollama_stream(prompt, model):
                answer += token

            sources = []
            for i, h in enumerate(hits, start=1):
                pl = h.payload or {}
                sources.append({
                    "rank": i,
                    "similarity": round(float(h.score), 4),
                    "source": pl.get("source", "unknown"),
                    "filename": pl.get("filename", "unknown"),
                    "chunk_id": pl.get("chunk_id", 0),
                    "preview": pl.get("text", "")[:300]
                })

            return {
                "answer": answer,
                "sources": sources,
                "mode": "rag"
            }
        else:
            # LLM mode
            system_prompt = req.system_prompt or (
                "당신은 도움이 되는 AI 어시스턴트입니다." if req.language == "ko"
                else "You are a helpful AI assistant."
            )

            answer = ""
            async for token in call_ollama_stream(req.question, model, system_prompt):
                answer += token

            return {
                "answer": answer,
                "sources": [],
                "mode": "llm"
            }
    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Query failed: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/search")
async def search_v1(
    query: str = Body(..., embed=True),
    top_k: int = Body(default=5, embed=True)
):
    """Vector search endpoint (API v1)"""
    try:
        client = get_qdrant()
        collections = [c.name for c in client.get_collections().collections]

        if COLLECTION_NAME not in collections:
            return {"results": [], "message": "No collection found"}

        qvec = embed_texts_batch([query])[0]
        hits = client.search(
            collection_name=COLLECTION_NAME,
            query_vector=qvec,
            limit=top_k
        )

        results = []
        for h in hits:
            pl = h.payload or {}
            results.append({
                "score": round(float(h.score), 4),
                "source": pl.get("source", "unknown"),
                "filename": pl.get("filename", "unknown"),
                "chunk_id": pl.get("chunk_id", 0),
                "text": pl.get("text", "")[:500]
            })

        return {"results": results, "count": len(results)}
    except Exception as e:
        log.error(f"Search failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/query/stream")
async def query_stream_v1(req: QueryRequest = Body(...)):
    """Streaming query endpoint (API v1)"""
    return await query_stream(req)

@app.post("/api/v1/upload")
async def upload_v1(files: List[UploadFile] = File(...)):
    """File upload endpoint (API v1)"""
    return await upload_files(files)

@app.get("/api/v1/collections")
async def collections_v1():
    """Get collections info (API v1)"""
    try:
        client = get_qdrant()
        collections = client.get_collections().collections

        result = []
        for col in collections:
            info = client.get_collection(col.name)
            result.append({
                "name": col.name,
                "vectors_count": info.vectors_count or 0,
                "points_count": info.points_count or 0
            })

        return {"collections": result, "count": len(result)}
    except Exception as e:
        log.error(f"Collections query failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/config")
def get_config():
    return {
        "ollama": {
            "host": OLLAMA_HOST,
            "port": OLLAMA_PORT,
            "model": OLLAMA_MODEL,
            "available": _ollama_available,
            "models": _available_models
        },
        "storage": {
            "data_dir": str(CURRENT_DATA_DIR),
            "exports_dir": str(EXPORT_DIR),
            "qdrant_mode": "embedded" if QDRANT_LOCAL else f"remote:{QDRANT_HOST}:{QDRANT_PORT}",
        },
        "chunk": {"size": CHUNK_SIZE, "overlap": CHUNK_OVERLAP},
        "features": {
            "rag_mode": True,
            "llm_mode": True,
            "multi_model": len(_available_models) > 1,
            "document_upload": True,
            "conversation_export": True
        }
    }

@app.post("/config/storage")
async def set_storage(cfg: StorageConfig):
    global CURRENT_DATA_DIR
    try:
        new_path = Path(cfg.path).expanduser().resolve()
        new_path.mkdir(parents=True, exist_ok=True)
        testfile = new_path / f".write_test_{uuid.uuid4().hex}"
        testfile.write_text("ok")
        testfile.unlink()
        CURRENT_DATA_DIR = new_path
        log.info(f"Storage path set to: {CURRENT_DATA_DIR}")
        return {"status": "success", "data_dir": str(CURRENT_DATA_DIR)}
    except Exception as e:
        log.error(f"Failed to set storage path: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid path: {str(e)}")


@app.post("/ingest", response_model=IngestResponse)
async def ingest_all():
    return await ingest_internal()

async def ingest_internal(saved_only: Optional[List[Path]] = None) -> IngestResponse:
    try:
        if saved_only:
            files = saved_only
        else:
            allowed = {".txt", ".md", ".pdf", ".docx", ".csv", ".xlsx"}
            files = [
                p for p in CURRENT_DATA_DIR.rglob("*")
                if p.is_file() and p.suffix.lower() in allowed
            ]

        if not files:
            return IngestResponse(files_indexed=0, chunks_indexed=0, status="no_files")

        em = get_embedding_model()
        probe_vec = [v for v in em.embed(["test"], batch_size=1)][0]
        vec_dim = len(probe_vec)
        ensure_collection(vec_dim)

        client = get_qdrant()

        all_texts = []
        all_metadata = []
        files_set = set()

        for fp in files:
            try:
                text = read_file_text(fp)
                if not text:
                    continue

                files_set.add(str(fp))
                chunk_id = 0

                for chunk in chunk_text(text):
                    all_texts.append(chunk)
                    all_metadata.append({
                        "source": str(fp),
                        "filename": fp.name,
                        "chunk_id": chunk_id,
                        "text": chunk,
                        "indexed_at": datetime.utcnow().isoformat()
                    })
                    chunk_id += 1
            except Exception as e:
                log.error(f"Failed to process {fp}: {e}")
                continue

        if not all_texts:
            return IngestResponse(files_indexed=0, chunks_indexed=0, status="no_content")

        vectors = embed_texts_batch(all_texts)

        points_data = []
        base_id = int(time.time() * 1000000)

        for i, (vec, meta) in enumerate(zip(vectors, all_metadata)):
            points_data.append({
                "id": base_id + i,
                "vector": vec,
                "payload": meta
            })

        batch_size = UPSERT_BATCH
        for i in range(0, len(points_data), batch_size):
            batch = points_data[i:i+batch_size]
            client.upsert(
                collection_name=COLLECTION_NAME,
                points=batch,
                wait=True
            )

        log.info(f"Indexed {len(files_set)} files, {len(all_texts)} chunks")

        return IngestResponse(
            files_indexed=len(files_set),
            chunks_indexed=len(all_texts),
            status="success"
        )

    except Exception as e:
        log.error(f"Ingest failed: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Indexing failed: {str(e)}")

@app.post("/query_stream")
async def query_stream(req: QueryRequest = Body(...)):
    try:
        async def generate():
            try:
                model = req.model or OLLAMA_MODEL

                # RAG 모드
                if req.mode == "rag":
                    client = get_qdrant()
                    collections = [c.name for c in client.get_collections().collections]

                    if COLLECTION_NAME not in collections:
                        yield json.dumps({"event": "error", "message": "No documents indexed"}) + "\n"
                        return

                    yield json.dumps({"event": "progress", "stage": "embedding", "pct": 10}) + "\n"
                    qvec = embed_texts_batch([req.question])[0]

                    yield json.dumps({"event": "progress", "stage": "searching", "pct": 30}) + "\n"

                    # 선택된 문서가 있으면 필터 적용
                    search_params = {
                        "collection_name": COLLECTION_NAME,
                        "query_vector": qvec,
                        "limit": req.top_k,
                    }

                    if req.selected_sources and len(req.selected_sources) > 0:
                        from qdrant_client.models import Filter, FieldCondition, MatchAny
                        search_params["query_filter"] = Filter(
                            must=[
                                FieldCondition(
                                    key="source",
                                    match=MatchAny(any=req.selected_sources)
                                )
                            ]
                        )
                        log.info(f"Filtering by {len(req.selected_sources)} selected sources")

                    hits = client.search(**search_params)

                    sources = []
                    for i, h in enumerate(hits, start=1):
                        pl = h.payload or {}
                        sources.append({
                            "rank": i,
                            "similarity": round(float(h.score), 4),
                            "source": pl.get("source", "unknown"),
                            "filename": pl.get("filename", "unknown"),
                            "chunk_id": pl.get("chunk_id", 0),
                            "preview": pl.get("text", "")[:300]
                        })

                    yield json.dumps({"event": "sources", "items": sources}) + "\n"
                    yield json.dumps({"event": "progress", "stage": "generating", "pct": 50}) + "\n"

                    contexts = [{"payload": h.payload, "score": float(h.score)} for h in hits]
                    prompt = build_rag_prompt(req.question, contexts, req.language)

                    async for token in call_ollama_stream(prompt, model):
                        yield json.dumps({"event": "token", "text": token}) + "\n"

                # 일반 LLM 모드
                else:
                    yield json.dumps({"event": "progress", "stage": "generating", "pct": 10}) + "\n"

                    system_prompt = req.system_prompt or (
                        "당신은 도움이 되는 AI 어시스턴트입니다." if req.language == "ko"
                        else "You are a helpful AI assistant."
                    )

                    async for token in call_ollama_stream(req.question, model, system_prompt):
                        yield json.dumps({"event": "token", "text": token}) + "\n"

                yield json.dumps({"event": "done", "pct": 100}) + "\n"

            except Exception as e:
                log.error(f"Query stream error: {e}\n{traceback.format_exc()}")
                yield json.dumps({"event": "error", "message": str(e)}) + "\n"

        return StreamingResponse(generate(), media_type="application/x-ndjson")

    except Exception as e:
        log.error(f"Query stream failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/vectors")
def vectors_summary():
    try:
        client = get_qdrant()
        collections = [c.name for c in client.get_collections().collections]

        if COLLECTION_NAME not in collections:
            return {"collection": COLLECTION_NAME, "total_points": 0, "sources": []}

        by_src: Dict[str, int] = {}
        total = 0

        scroll_result = client.scroll(
            collection_name=COLLECTION_NAME,
            limit=10000,
            with_payload=True,
        )

        for pt in scroll_result[0]:
            src = (pt.payload or {}).get("source", "unknown")
            by_src[src] = by_src.get(src, 0) + 1
            total += 1

        items = [{"source": k, "points": v} for k, v in sorted(by_src.items(), key=lambda x: (-x[1], x[0]))]

        return {"collection": COLLECTION_NAME, "total_points": total, "sources": items}

    except Exception as e:
        log.error(f"Vector summary failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/upload")
async def upload_files(files: List[UploadFile] = File(...)):
    """파일 업로드 및 벡터 색인"""
    try:
        if not files:
            raise HTTPException(status_code=400, detail="파일이 제공되지 않았습니다")

        # 파일 저장 및 처리
        uploaded_files = []
        for file in files:
            if not file.filename:
                continue

            # 파일 저장
            file_path = CURRENT_DATA_DIR / file.filename
            content = await file.read()

            with open(file_path, 'wb') as f:
                f.write(content)

            log.info(f"File saved: {file.filename} ({len(content)} bytes)")
            uploaded_files.append(str(file_path))

        if not uploaded_files:
            raise HTTPException(status_code=400, detail="유효한 파일이 없습니다")

        # 파일 읽기 및 청킹
        all_chunks = []
        log.info(f"Processing {len(uploaded_files)} uploaded files...")

        for fpath_str in uploaded_files:
            fpath = Path(fpath_str)
            log.info(f"Processing file: {fpath}")

            try:
                # 텍스트 추출
                text = read_file_text(fpath)
                log.info(f"Extracted text length: {len(text) if text else 0} chars")

                if not text or not text.strip():
                    log.warning(f"Empty or whitespace-only text from {fpath.name}")
                    log.warning(f"Text content preview: '{text[:100] if text else 'None'}'")
                    continue

                log.info(f"Text preview (first 200 chars): {text[:200]}")

                # 청킹
                chunks_list = list(chunk_text(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP))
                log.info(f"File {fpath.name}: {len(chunks_list)} chunks created")

                if chunks_list:
                    log.info(f"First chunk preview: {chunks_list[0][:100]}")

                for i, chunk_content in enumerate(chunks_list):
                    all_chunks.append({
                        "text": chunk_content,
                        "source": str(fpath),
                        "filename": fpath.name,
                        "chunk_id": i
                    })

            except Exception as e:
                log.error(f"Failed to process {fpath.name}: {e}")
                import traceback
                log.error(traceback.format_exc())
                continue

        log.info(f"Total chunks created: {len(all_chunks)}")

        if not all_chunks:
            log.error("No chunks created from any file!")
            raise HTTPException(status_code=400, detail="처리 가능한 텍스트가 없습니다")

        # 임베딩 생성
        log.info(f"Generating embeddings for {len(all_chunks)} chunks...")
        texts = [c["text"] for c in all_chunks]

        em = get_embedding_model()
        embeddings = []

        for batch_start in range(0, len(texts), FASTEMBED_BATCH):
            batch_end = min(batch_start + FASTEMBED_BATCH, len(texts))
            batch_texts = texts[batch_start:batch_end]

            for vec in em.embed(batch_texts, batch_size=len(batch_texts)):
                embeddings.append(vec.tolist())

        if len(embeddings) != len(all_chunks):
            raise HTTPException(
                status_code=500,
                detail=f"임베딩 개수 불일치: {len(embeddings)} vs {len(all_chunks)}"
            )

        # Qdrant 컬렉션 준비
        client = get_qdrant()
        collections = [c.name for c in client.get_collections().collections]

        if COLLECTION_NAME not in collections:
            dim = len(embeddings[0])
            client.create_collection(
                collection_name=COLLECTION_NAME,
                vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
            )
            log.info(f"Created collection: {COLLECTION_NAME}")

        # 벡터 업로드
        from qdrant_client.models import PointStruct

        points = []
        for idx, (chunk, vec) in enumerate(zip(all_chunks, embeddings)):
            points.append(
                PointStruct(
                    id=str(uuid.uuid4()),
                    vector=vec,
                    payload={
                        "text": chunk["text"],
                        "source": chunk["source"],
                        "filename": chunk["filename"],
                        "chunk_id": chunk["chunk_id"]
                    }
                )
            )

        # 배치 업로드
        for batch_start in range(0, len(points), UPSERT_BATCH):
            batch_end = min(batch_start + UPSERT_BATCH, len(points))
            batch_points = points[batch_start:batch_end]

            client.upsert(
                collection_name=COLLECTION_NAME,
                points=batch_points
            )

        log.info(f"Uploaded {len(points)} vectors to Qdrant")

        return {
            "status": "success",
            "files": len(uploaded_files),
            "chunks": len(all_chunks),
            "vectors": len(points),
            "filenames": [Path(f).name for f in uploaded_files],
            "uploaded_paths": uploaded_files  # 전체 경로 반환 (자동 선택용)
        }

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Upload failed: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"업로드 실패: {str(e)}")

@app.delete("/vectors")
def vectors_delete(source: Optional[str] = Query(None), delete_all: bool = Query(False)):
    try:
        client = get_qdrant()
        collections = [c.name for c in client.get_collections().collections]

        if COLLECTION_NAME not in collections:
            return {"deleted": 0, "status": "no_collection"}

        if delete_all:
            em = get_embedding_model()
            probe_vec = [v for v in em.embed(["test"], batch_size=1)][0]
            dim = len(probe_vec)

            client.recreate_collection(
                collection_name=COLLECTION_NAME,
                vectors_config=VectorParams(size=dim, distance=Distance.COSINE),
            )
            log.info("Deleted all vectors")
            return {"deleted": "all", "status": "success"}

        if not source:
            raise HTTPException(status_code=400, detail="source or delete_all required")

        flt = Filter(must=[FieldCondition(key="source", match=MatchValue(value=source))])
        client.delete(collection_name=COLLECTION_NAME, points_selector=FilterSelector(filter=flt))

        log.info(f"Deleted vectors for source: {source}")
        return {"deleted": "by_source", "source": source, "status": "success"}

    except Exception as e:
        log.error(f"Vector deletion failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/conversation/export")
async def export_conversation(conv: ConversationExport):
    """대화 내보내기"""
    try:
        export_file = CHAT_HISTORY_DIR / f"{conv.conversation_id}.json"
        export_file.write_text(json.dumps(conv.dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        return {"status": "success", "file": str(export_file)}
    except Exception as e:
        log.error(f"Export failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/document/source")
async def get_document_source(
    path: str = Query(..., description="문서 파일 경로"),
    chunk_id: Optional[int] = Query(None, description="청크 ID (하이라이트할 청크)")
):
    """원본 문서 내용 가져오기 - PDF, DOCX, TXT, MD, CSV, XLSX 지원"""
    try:
        log.info(f"📄 원본 문서 요청: {path}, chunk_id: {chunk_id}")

        # 파일 경로 검증
        file_path = Path(path)

        if not file_path.exists():
            log.warning(f"파일을 찾을 수 없음: {path}")
            raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

        if not file_path.is_file():
            log.warning(f"파일이 아님: {path}")
            raise HTTPException(status_code=400, detail="올바른 파일이 아닙니다")

        # read_file_text 함수를 사용하여 모든 파일 형식 지원
        content = read_file_text(file_path)

        if not content:
            log.warning(f"파일 내용이 비어있음: {path}")
            raise HTTPException(status_code=400, detail="파일 내용을 읽을 수 없습니다")

        log.info(f"✅ 원본 문서 로드 완료: {len(content)} 문자")

        # 청크 정보 가져오기 (chunk_id가 제공된 경우)
        chunk_info = None
        if chunk_id is not None:
            try:
                # Qdrant에서 해당 청크 검색
                client = get_qdrant()
                collections = [c.name for c in client.get_collections().collections]

                if COLLECTION_NAME in collections:
                    # source와 chunk_id로 필터링하여 검색
                    scroll_result = client.scroll(
                        collection_name=COLLECTION_NAME,
                        scroll_filter=Filter(
                            must=[
                                FieldCondition(key="source", match=MatchValue(value=str(file_path))),
                                FieldCondition(key="chunk_id", match=MatchValue(value=chunk_id))
                            ]
                        ),
                        limit=1,
                        with_payload=True
                    )

                    if scroll_result[0]:
                        chunk_payload = scroll_result[0][0].payload
                        chunk_text = chunk_payload.get("text", "")

                        # 원본 문서에서 청크 위치 찾기
                        chunk_start = content.find(chunk_text[:100])  # 청크의 첫 100자로 검색
                        if chunk_start != -1:
                            chunk_info = {
                                "chunk_id": chunk_id,
                                "text": chunk_text,
                                "start_pos": chunk_start,
                                "end_pos": chunk_start + len(chunk_text)
                            }
                            log.info(f"✅ 청크 위치 찾음: {chunk_start} - {chunk_start + len(chunk_text)}")
            except Exception as e:
                log.warning(f"청크 정보 가져오기 실패: {e}")

        return {
            "status": "success",
            "path": str(file_path),
            "filename": file_path.name,
            "file_type": file_path.suffix.lower(),
            "size": len(content),
            "content": content,
            "chunk_info": chunk_info
        }

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"원본 문서 로드 오류: {e}")
        log.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"원본 문서를 불러올 수 없습니다: {str(e)}")

# ==================== 실행 ====================
if __name__ == "__main__":
    import uvicorn
    log.info("Starting Private RAG API Server - Fabrix Edition...")
    uvicorn.run(
        "main:app",
        host="127.0.0.1",
        port=8000,
        reload=False,
        log_level="info"
    )
